"""
[SUB-ROTINA] EXCEL EXPORTER
Usa apenas 'openpyxl' e 'sqlite3'.
Otimizado para memória: Itera sobre o cursor sem carregar tudo em lista.
"""
import sys
import argparse
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def export_excel(cursor: sqlite3.Cursor, out_path: str) -> None:
    """
    Gera o Excel iterando diretamente sobre o cursor (Lazy Load).
    """
    # 1. Cria Workbook e Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    # 2. Cabeçalhos
    if cursor.description:
        headers = [desc[0] for desc in cursor.description]
    else:
        headers = []
    
    ws.append(headers)
    
    # --- ESTILIZAÇÃO DO CABEÇALHO ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[1].height = 30 
    ws.freeze_panes = "A2"
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 3. Dados (Streaming)
    # Lista auxiliar para calcular largura das colunas on-the-fly
    # Começa com o tamanho do header
    col_widths = [len(str(h)) + 2 for h in headers]
    
    row_count = 0
    
    # ITERAÇÃO DIRETA NO CURSOR (Memória Baixa)
    for row in cursor:
        ws.append(row)
        row_count += 1
        current_row_idx = ws.max_row
        
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row_idx, column=col_idx)
            
            # --- Formatação ---
            if isinstance(value, int) and not isinstance(value, bool):
                cell.number_format = '0' 
            elif isinstance(value, float):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
            
            # --- Cálculo de Largura (Otimizado) ---
            # Só atualiza se for maior que o atual e menor que o limite (50)
            if value is not None:
                curr_len = len(str(value))
                if curr_len > col_widths[col_idx-1]:
                    col_widths[col_idx-1] = curr_len

    # 4. Aplica larguras finais e filtros
    ws.auto_filter.ref = ws.dimensions
    
    for i, column_cells in enumerate(ws.columns):
        # Limita a 60 chars para não quebrar layout
        final_width = min(col_widths[i] + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = final_width

    wb.save(out_path)
    print(f"[EXCEL] 💾 Salvo: {Path(out_path).name} ({row_count} linhas)")

# --- MODO STANDALONE (Para testes manuais) ---
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True)
    parser.add_argument("--sql", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    try:
        conn = sqlite3.connect(args.db)
        cursor = conn.cursor()
        cursor.execute(args.sql)
        
        export_excel(cursor, args.out)
        
        conn.close()
        sys.exit(0)
    except Exception as e:
        print(f"[ERRO EXCEL] {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()